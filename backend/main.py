"""
GRACE Prototype — FastAPI Backend
All REST endpoints — /api/v1/
"""
import json
import io
import re
from pathlib import Path
from typing import Optional
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import fitz  # PyMuPDF
import docx2txt

from modules.database import (
    init_db, register_document, get_document, list_documents,
    create_run, complete_run, get_run, list_runs,
    save_findings, list_findings, get_finding, update_operational_status,
    get_kpi_summary, save_output_document, list_output_documents,
    count_mappings,
    # Phase 1: Risks
    create_risk, get_risk, list_risks, update_risk, delete_risk,
    # Phase 1: Vendors
    create_vendor, get_vendor, list_vendors, update_vendor,
    save_vendor_assessment, delete_vendor, DEFAULT_VENDOR_QUESTIONS,
    # Phase 1: Policies + Acknowledgments
    create_policy, get_policy, list_policies, update_policy,
    assign_policy, list_policy_assignments, acknowledge_assignment,
    # Phase 1: Incidents
    create_incident, get_incident, list_incidents, update_incident,
    delete_incident,
    # Ask GRACE: chat persistence
    create_chat_session, get_chat_session, list_chat_sessions,
    update_chat_session_title, delete_chat_session,
    append_chat_message, list_chat_messages,
)
from modules.grc_engine import (
    run_gap_analysis, generate_document, explain_control,
    list_supported_frameworks, load_framework,
    get_or_compute_mappings, generate_vendor_assessment_summary,
)

app = FastAPI(
    title="GRACE API",
    description="Governance, Risk, Assurance & Compliance Engine — Prototype v1.0",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # prototype only — restrict in production
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialise DB on startup
@app.on_event("startup")
def on_startup():
    init_db()


# ─── Health ──────────────────────────────────────────────────────────

@app.get("/health")
def health():
    """Real liveness/readiness probe.

    - 'ok'       : DB reachable AND Anthropic API key configured
    - 'degraded' : DB reachable but API key missing (read-only)
    - 'down'     : DB unreachable (returns HTTP 503)
    """
    import os
    from fastapi import HTTPException
    from modules.database import get_db

    try:
        conn = get_db()
        conn.execute("SELECT 1").fetchone()
        conn.close()
        db_ok = True
    except Exception:
        db_ok = False

    api_key_ok = bool(os.environ.get("ANTHROPIC_API_KEY", "").strip())

    if not db_ok:
        raise HTTPException(status_code=503, detail={
            "status": "down", "service": "GRACE API", "db": False,
        })

    status = "ok" if api_key_ok else "degraded"
    return {
        "status": status,
        "service": "GRACE API",
        "version": "1.0.0-prototype",
        "db": True,
        "api_key": api_key_ok,
    }


# ─── Admin (prototype only) ──────────────────────────────────────────

@app.post("/api/v1/admin/reset")
def reset_prototype_data():
    """Wipe all user-generated data — uploaded/generated documents,
    assessment runs, findings, gaps, translations and output documents.
    Framework catalogs and the schema itself are untouched. Prototype-
    only convenience endpoint, not safe for production use."""
    from modules.database import get_db
    conn = get_db()
    tables = [
        "finding_translations",
        "gaps",
        "findings",
        "audit_events",
        "assessment_runs",
        "output_documents",
        "documents",
        # Phase 1 GRC modules
        "risks",
        "vendors",
        "policy_assignments",
        "policies",
        "incidents",
    ]
    counts = {}
    for table in tables:
        try:
            row = conn.execute(f"SELECT COUNT(*) as n FROM {table}").fetchone()
            counts[table] = row["n"]
            conn.execute(f"DELETE FROM {table}")
        except Exception:
            counts[table] = None
    conn.commit()
    conn.close()
    return {"status": "reset_complete", "deleted": counts}


# ─── Frameworks ──────────────────────────────────────────────────────

@app.get("/api/v1/frameworks")
def get_frameworks():
    return {"frameworks": list_supported_frameworks()}


@app.get("/api/v1/frameworks/{framework_id}/controls")
def get_controls(framework_id: str):
    fw = load_framework(framework_id)
    if not fw:
        raise HTTPException(404, detail=f"Framework {framework_id} not found or not loaded")
    return {"framework_id": framework_id, "controls": fw.get("controls", [])}


@app.get("/api/v1/frameworks/{framework_id}/controls/{control_id}/explain")
def explain(framework_id: str, control_id: str, language: Optional[str] = "en"):
    explanation = explain_control(framework_id, control_id, language=language)
    return {"framework_id": framework_id, "control_id": control_id, "explanation": explanation}


# ─── Documents ───────────────────────────────────────────────────────

class DocumentText(BaseModel):
    title: str
    content: str
    owner: Optional[str] = "demo"
    business_unit: Optional[str] = "Security"


@app.post("/api/v1/documents/text", status_code=201)
def register_text_document(body: DocumentText):
    result = register_document(body.title, body.content, body.owner, body.business_unit)
    return result


@app.post("/api/v1/documents/upload", status_code=201)
async def upload_document(
    file: UploadFile = File(...),
    owner: str = Form("demo"),
    business_unit: str = Form("Security")
):
    content_bytes = await file.read()
    fname = file.filename or "document"
    ext = Path(fname).suffix.lower()

    if ext == ".pdf":
        with fitz.open(stream=content_bytes, filetype="pdf") as doc:
            text = "\n".join(page.get_text() for page in doc)
    elif ext in (".docx", ".doc"):
        text = docx2txt.process(io.BytesIO(content_bytes))
    elif ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            parts.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(cells):
                    parts.append("\t".join(cells))
        text = "\n".join(parts)
    elif ext == ".csv":
        import csv
        raw = content_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(raw))
        text = "\n".join("\t".join(row) for row in reader)
    else:
        text = content_bytes.decode("utf-8", errors="replace")

    if not text.strip():
        raise HTTPException(400, detail="Could not extract text from document")

    result = register_document(fname, text, owner, business_unit)
    result["character_count"] = len(text)
    return result


@app.get("/api/v1/documents")
def get_documents():
    return {"documents": list_documents()}


@app.get("/api/v1/documents/output")
def get_output_documents():
    return {"documents": list_output_documents()}


@app.get("/api/v1/documents/{document_id}")
def get_doc(document_id: str):
    doc = get_document(document_id)
    if not doc:
        raise HTTPException(404, detail="Document not found")
    return doc


# ─── Assessment ──────────────────────────────────────────────────────

class AssessmentRequest(BaseModel):
    document_id: str
    framework: str
    controls_scope: Optional[list] = None
    channel: Optional[str] = "web"
    language: Optional[str] = "en"


# In-memory run status store (lightweight for prototype)
_run_results: dict = {}


def _run_assessment_bg(run_id: str, document_id: str, framework: str,
                       controls_scope: list, language: str = "en"):
    """Background task: run assessment and persist results."""
    try:
        doc = get_document(document_id)
        if not doc:
            complete_run(run_id, "error", "Document not found")
            return

        def progress(msg):
            _run_results[run_id] = {"status": "running", "message": msg}

        result = run_gap_analysis(
            document_text=doc["content_text"],
            document_title=doc["title"],
            framework_id=framework,
            controls_scope=controls_scope,
            progress_callback=progress,
            language=language,
        )
        finding_ids = save_findings(run_id, document_id, result, language=language)
        complete_run(run_id, "completed")
        _run_results[run_id] = {
            "status": "completed",
            "result": result,
            "finding_ids": finding_ids
        }
    except Exception as e:
        complete_run(run_id, "error", str(e))
        _run_results[run_id] = {"status": "error", "error": str(e)}


@app.post("/api/v1/assessments/run", status_code=202)
def start_assessment(body: AssessmentRequest, background_tasks: BackgroundTasks):
    run_id = create_run(body.document_id, body.framework, body.controls_scope, body.channel)
    _run_results[run_id] = {"status": "pending"}
    background_tasks.add_task(
        _run_assessment_bg, run_id, body.document_id, body.framework,
        body.controls_scope, body.language
    )
    return {"run_id": run_id, "status": "pending", "message": "Assessment started"}


@app.get("/api/v1/assessments/{run_id}")
def poll_run(run_id: str):
    in_memory = _run_results.get(run_id)
    if in_memory:
        return in_memory
    db_run = get_run(run_id)
    if not db_run:
        raise HTTPException(404, detail="Run not found")
    return {"status": db_run["status"]}


@app.get("/api/v1/assessments")
def get_runs():
    return {"runs": list_runs()}


# ─── Synchronous assessment (for Streamlit polling UI) ───────────────

@app.post("/api/v1/assessments/run-sync")
def run_assessment_sync(body: AssessmentRequest):
    """Run assessment synchronously and return result immediately."""
    doc = get_document(body.document_id)
    if not doc:
        raise HTTPException(404, detail="Document not found")

    run_id = create_run(body.document_id, body.framework, body.controls_scope, body.channel)
    try:
        result = run_gap_analysis(
            document_text=doc["content_text"],
            document_title=doc["title"],
            framework_id=body.framework,
            controls_scope=body.controls_scope,
            language=body.language,
        )
        finding_ids = save_findings(run_id, body.document_id, result, language=body.language or "en")
        complete_run(run_id, "completed")
        return {"run_id": run_id, "status": "completed", "result": result, "finding_ids": finding_ids}
    except Exception as e:
        complete_run(run_id, "error", str(e))
        raise HTTPException(500, detail=str(e))


# ─── Ask GRACE: unified natural-language workspace ───────────────────
#
# Single entry point that routes the request to the right downstream
# pipeline. Keeps the old endpoints intact (callers like the legacy
# Assessment widget can still hit them), but the new Ask GRACE panel
# in the UI talks ONLY to /api/v1/ask. Intent classification is a
# straightforward keyword heuristic today — well-isolated so we can
# swap it for a Claude-based router later without touching the
# dispatch logic.

class AskRequest(BaseModel):
    query: str
    document_ids: Optional[list] = None
    framework_id: Optional[str] = None
    language: Optional[str] = "en"


_INTENT_KEYWORDS = {
    "explain":         ["explain ", "spiega ", "what is ", "cosa è ", "cos'è "],
    "query_findings":  ["finding", "registr", "open", "aperti", "critical", "critici",
                        "list ", "show me", "mostra", "summari", "riassum"],
    # NOTE: the historical `analyze_new` intent was removed. Structured
    # assessments now live exclusively on /api/v1/assessments/run-sync,
    # driven by the dedicated Gap Analysis wizard. Ask GRACE is purely
    # conversational and never persists findings.
}


def classify_intent(query: str, has_docs: bool, has_framework: bool) -> str:
    """Lightweight intent router for Ask GRACE.

    Priority:
      1. explain       — query starts with 'explain X', 'what is X' …
      2. query_findings — query mentions findings/registry/critical/etc.
      3. document_qa   — everything else (with or without attached docs)

    The `has_docs` / `has_framework` flags are kept in the signature so the
    chat layer can pass them through if we want to specialise routing
    later, but they no longer trigger a persisted assessment.
    """
    q = (query or "").strip().lower()
    for kw in _INTENT_KEYWORDS["explain"]:
        if q.startswith(kw):
            return "explain"
    for kw in _INTENT_KEYWORDS["query_findings"]:
        if kw in q:
            return "query_findings"
    return "document_qa"


def _document_qa_response(query: str, framework_id: Optional[str], language: str,
                          document_ids: Optional[list] = None) -> str:
    """Generic Claude Q&A about GRC. Lightweight system prompt that
    keeps GRACE on-topic without long-context overhead. When
    `document_ids` are supplied, their text is injected into the user
    message so Claude can reason over uploaded/pasted context even
    without a framework selection (e.g. cross-document mapping)."""
    from modules.grc_engine import get_client, _language_instruction
    system = (
        "You are GRACE, an AI GRC (Governance, Risk, Assurance & Compliance) "
        "analyst. Reply to the user's question concisely, with practical "
        "framing. Cite framework articles or control IDs when relevant. If "
        "the question is outside GRC scope, say so briefly and redirect to "
        "what GRACE can help with. When documents are provided in the "
        "context, ground every claim in them and quote the relevant "
        "passages verbatim."
        + _language_instruction(language, mode="prose")
    )

    # Build the user message, prepending any attached document content.
    # Cap each document at ~40k chars to stay well under the model's
    # context window when several large files are attached.
    doc_blocks = []
    for did in (document_ids or []):
        d = get_document(did)
        if d and d.get("content_text"):
            body_text = d["content_text"][:40000]
            doc_blocks.append(
                f"<document title=\"{d.get('title', 'Untitled')}\">\n"
                f"{body_text}\n"
                f"</document>"
            )

    parts = []
    if doc_blocks:
        parts.append(
            "The user has attached the following document(s) as context. "
            "Use them as your primary source of truth:"
        )
        parts.extend(doc_blocks)
        parts.append("---")
    if framework_id:
        parts.append(f"(Context: user's working framework is {framework_id})")
    parts.append(f"User question: {query}")
    user_msg = "\n\n".join(parts)

    # Sonnet handles the larger context windows that document Q&A needs;
    # Haiku stays the default only for the no-document path.
    model = "claude-sonnet-4-6" if doc_blocks else "claude-haiku-4-5-20251001"
    try:
        msg = get_client().messages.create(
            model=model,
            max_tokens=1500 if doc_blocks else 900,
            system=system,
            messages=[{"role": "user", "content": user_msg}],
        )
        return msg.content[0].text
    except Exception as e:
        return f"_Couldn't reach the model: {e}_"


def _findings_summary_response(query: str, language: str) -> tuple[str, list]:
    """Summarise the findings table in the user's words via Claude.
    Returns (markdown, citations)."""
    findings = list_findings(limit=50)
    if not findings:
        return ("_No findings in the registry yet — run a Gap Analysis to populate it._", [])
    # Build a compact tabular context for Claude
    rows = []
    for f in findings[:30]:
        rows.append(
            f"- [{f.get('severity','?').upper()}] {f.get('framework','')} "
            f"{f.get('control_id','')} · {f.get('control_title','')[:80]} "
            f"(status: {f.get('compliance_status','?')}, op: {f.get('operational_status','?')})"
        )
    from modules.grc_engine import get_client, _language_instruction
    system = (
        "You are GRACE, an AI GRC analyst. You will be given a list of "
        "findings from the user's registry. Answer the user's question by "
        "summarising / filtering / explaining those findings. Be concise, "
        "highlight critical and high severity items first, and reference "
        "control IDs verbatim."
        + _language_instruction(language, mode="prose")
    )
    user_msg = (
        f"User question: {query}\n\n"
        f"Findings registry ({len(findings)} total, showing top 30):\n"
        + "\n".join(rows)
    )
    try:
        msg = get_client().messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=900,
            system=system,
            messages=[{"role": "user", "content": user_msg}],
        )
        text = msg.content[0].text
    except Exception as e:
        text = f"_Couldn't reach the model: {e}_"
    citations = [
        {"type": "finding", "id": f["finding_id"],
         "label": f"{f.get('framework','')} {f.get('control_id','')}"}
        for f in findings[:8]
    ]
    return (text, citations)


def _dispatch_ask(query: str, document_ids: Optional[list],
                  framework_id: Optional[str], language: str) -> dict:
    """Shared Ask-GRACE dispatch logic — used by /api/v1/ask AND by the
    chat-message POST. Returns the same envelope shape as before, minus
    the removed `analyze_new` branch.

    Envelope:
      {
        "intent":        "explain" | "query_findings" | "document_qa",
        "response_type": "explanation" | "findings_qa" | "qa",
        "response_text": "<markdown>",
        "citations":     [{"type": "...", "id": "...", "label": "..."}]
      }
    """
    document_ids = document_ids or []
    has_docs = bool(document_ids)
    has_fw   = bool(framework_id)
    intent   = classify_intent(query or "", has_docs, has_fw)
    language = language or "en"

    if intent == "explain":
        # Try to parse 'explain CONTROL_ID' or 'spiega CONTROL_ID'.
        # Fall back to document_qa if no recognisable control ID.
        import re as _re
        m = _re.search(r"([A-Z]+\.?[A-Z0-9.\-]+|Art\.\d+(?:\.\d+)*\.?[a-z]?)", query or "")
        if m and framework_id:
            ctrl_id = m.group(1)
            from modules.grc_engine import explain_control
            text = explain_control(framework_id, ctrl_id, language=language)
            return {
                "intent":        intent,
                "response_type": "explanation",
                "response_text": text,
                "citations":     [{"type": "control", "id": ctrl_id,
                                   "label": f"{framework_id} · {ctrl_id}"}],
            }
        intent = "document_qa"

    if intent == "query_findings":
        text, citations = _findings_summary_response(query, language)
        return {
            "intent":        intent,
            "response_type": "findings_qa",
            "response_text": text,
            "citations":     citations,
        }

    # document_qa — attached documents (with or without a framework) flow
    # through here as document-grounded Q&A. NEVER persists findings;
    # users wanting a structured assessment go through the Gap Analysis
    # wizard at /api/v1/assessments/run-sync.
    text = _document_qa_response(query, framework_id, language,
                                 document_ids=document_ids)
    titles = []
    for did in document_ids:
        d = get_document(did)
        if d:
            titles.append(d["title"])
    return {
        "intent":        "document_qa",
        "response_type": "qa",
        "response_text": text,
        "citations":     [{"type": "document", "id": did, "label": ttl}
                          for did, ttl in zip(document_ids, titles)],
    }


@app.post("/api/v1/ask")
def ask_grace(body: AskRequest):
    """Unified Ask-GRACE dispatcher. Kept for backwards compatibility with
    any caller that still talks to /api/v1/ask directly; the new chat UI
    routes through /api/v1/chat/sessions/{id}/messages, which calls the
    same `_dispatch_ask` helper.

    NOTE: as of the Ask GRACE / Gap Analysis split, this endpoint no
    longer triggers a persisted gap analysis even when document_ids +
    framework_id are both supplied — that path moved to the dedicated
    /api/v1/assessments/run-sync endpoint.
    """
    return _dispatch_ask(
        query=body.query or "",
        document_ids=body.document_ids,
        framework_id=body.framework_id,
        language=body.language or "en",
    )


# ─── Chat ─────────────────────────────────────────────────────────────
#
# Persistent chat sessions for the Ask GRACE page. Each session is a
# thread of messages; the assistant reply is produced by `_dispatch_ask`,
# so chat and the legacy /api/v1/ask endpoint share intent routing.

class ChatSessionCreate(BaseModel):
    title: Optional[str] = None


class ChatSessionRename(BaseModel):
    title: str


class ChatMessageCreate(BaseModel):
    query: str
    document_ids: Optional[list] = None
    framework_id: Optional[str] = None
    language: Optional[str] = "en"


def _autotitle_from_query(q: str) -> str:
    """First 60 chars of the user's opening message, cleaned of newlines."""
    s = (q or "").strip().replace("\n", " ")
    if len(s) <= 60:
        return s or "Untitled"
    return s[:57].rstrip() + "…"


@app.post("/api/v1/chat/sessions")
def chat_create_session(body: ChatSessionCreate):
    return create_chat_session(user_id="demo", title=body.title)


@app.get("/api/v1/chat/sessions")
def chat_list_sessions(limit: int = 50):
    return {"sessions": list_chat_sessions(user_id="demo", limit=limit)}


@app.get("/api/v1/chat/sessions/{session_id}")
def chat_get_session(session_id: str):
    sess = get_chat_session(session_id)
    if not sess:
        raise HTTPException(404, detail="Session not found")
    sess["messages"] = list_chat_messages(session_id)
    return sess


@app.patch("/api/v1/chat/sessions/{session_id}")
def chat_rename_session(session_id: str, body: ChatSessionRename):
    sess = get_chat_session(session_id)
    if not sess:
        raise HTTPException(404, detail="Session not found")
    update_chat_session_title(session_id, body.title)
    return get_chat_session(session_id)


@app.delete("/api/v1/chat/sessions/{session_id}")
def chat_delete_session(session_id: str):
    sess = get_chat_session(session_id)
    if not sess:
        raise HTTPException(404, detail="Session not found")
    delete_chat_session(session_id)
    return {"status": "deleted", "session_id": session_id}


@app.get("/api/v1/chat/sessions/{session_id}/messages")
def chat_get_messages(session_id: str, limit: int = 200):
    sess = get_chat_session(session_id)
    if not sess:
        raise HTTPException(404, detail="Session not found")
    return {"messages": list_chat_messages(session_id, limit=limit)}


@app.post("/api/v1/chat/sessions/{session_id}/messages")
def chat_post_message(session_id: str, body: ChatMessageCreate):
    sess = get_chat_session(session_id)
    if not sess:
        raise HTTPException(404, detail="Session not found")

    user_msg = append_chat_message(
        session_id, role="user", content=body.query or "",
        document_ids=body.document_ids,
        framework_id=body.framework_id,
    )

    # If this is the FIRST message and the session has no title yet,
    # auto-title from the opening query so the sidebar list is readable.
    if not sess.get("title"):
        update_chat_session_title(session_id, _autotitle_from_query(body.query or ""))

    resp = _dispatch_ask(
        query=body.query or "",
        document_ids=body.document_ids,
        framework_id=body.framework_id,
        language=body.language or "en",
    )

    asst_msg = append_chat_message(
        session_id, role="assistant",
        content=resp.get("response_text", ""),
        document_ids=body.document_ids,
        framework_id=body.framework_id,
        citations=resp.get("citations"),
        intent=resp.get("intent"),
        response_type=resp.get("response_type"),
    )
    return {"user_message": user_msg, "assistant_message": asst_msg}


# ─── Findings ────────────────────────────────────────────────────────

@app.get("/api/v1/findings")
def get_findings(framework: Optional[str] = None, status: Optional[str] = None,
                 operational_status: Optional[str] = None, limit: int = 100,
                 language: Optional[str] = None):
    """Return findings, with user-facing fields lazily translated to
    `language` when it differs from the finding's stored generation
    language. Translations are cached in the finding_translations table
    so each (finding × target_lang) pair only hits Claude once."""
    from modules.database import (
        get_finding_translation, save_finding_translation,
    )
    from modules.grc_engine import translate_finding_fields

    findings = list_findings(framework, status, limit, operational_status)

    # Decorate every finding with `cross_framework_count` — purely from
    # the cache, no Claude call here. The badge in the Registry header
    # uses this; first time a finding is opened the count may be 0
    # (lazy on detail expand), but on subsequent loads it shows the real
    # number. Invariant: this MUST be cheap — list_findings is on the
    # hot path of the registry page.
    for f in findings:
        try:
            f["cross_framework_count"] = count_mappings(
                f.get("framework", ""), f.get("control_id", ""))
        except Exception:
            f["cross_framework_count"] = 0

    if not language or language not in ("en", "it"):
        return {"findings": findings}

    # We deliberately do NOT short-circuit on (stored_language == requested
    # language). The stored language column reflects the language passed
    # to the run, but Claude's output can still come back in a different
    # language (e.g. if the source document was Italian and the prompt
    # only weakly steered output to English). For every (finding ×
    # ui_language) we serve from cache or ask Claude.
    #
    # The first view per pair triggers one Haiku round-trip per finding.
    # We run those Claude calls in PARALLEL via a ThreadPoolExecutor so a
    # 10-finding page costs ~one Haiku-call wall-clock, not ten — the
    # frontend no longer freezes for 30 s on the first Registry visit.
    import concurrent.futures as _cf
    from threading import Lock

    pending = []        # findings that need a Claude call
    for f in findings:
        fid = f["finding_id"]
        cached = get_finding_translation(fid, language)
        if cached:
            f["description"]          = cached["description"]
            f["recommended_action"]   = cached["recommended_action"]
            f["regulatory_reference"] = cached["regulatory_reference"]
            if cached.get("control_title"):
                f["control_title"]    = cached["control_title"]
            f["language"]             = language
        else:
            pending.append(f)

    if pending:
        def _translate_one(f):
            return f, translate_finding_fields(
                f.get("description", ""),
                f.get("recommended_action", ""),
                f.get("regulatory_reference", ""),
                language,
                control_title=f.get("control_title", ""),
            )

        with _cf.ThreadPoolExecutor(max_workers=min(8, len(pending))) as ex:
            for f, translated in ex.map(_translate_one, pending):
                # Apply the translated content in-flight regardless of
                # outcome — when ok=False the originals stay in place,
                # which is the safest fallback for the UI.
                f["description"]          = translated["description"]
                f["recommended_action"]   = translated["recommended_action"]
                f["regulatory_reference"] = translated["regulatory_reference"]
                if translated.get("control_title"):
                    f["control_title"]    = translated["control_title"]
                f["language"]             = language
                # CRITICAL: only persist to the translation cache when
                # the call actually succeeded.
                if translated.get("ok"):
                    save_finding_translation(
                        f["finding_id"], language,
                        translated["description"],
                        translated["recommended_action"],
                        translated["regulatory_reference"],
                        success=True,
                        control_title=translated.get("control_title", ""),
                    )

    return {"findings": findings}


class StatusUpdate(BaseModel):
    operational_status: str
    actor: Optional[str] = "demo_user"


@app.patch("/api/v1/findings/{finding_id}/status")
def patch_finding_status(finding_id: str, body: StatusUpdate):
    valid = {"new","acknowledged","in_progress","resolved","accepted_risk","closed","dismissed"}
    if body.operational_status not in valid:
        raise HTTPException(400, detail=f"Invalid status. Must be one of: {valid}")
    update_operational_status(finding_id, body.operational_status, body.actor)
    return {"finding_id": finding_id, "operational_status": body.operational_status}


@app.get("/api/v1/findings/{finding_id}/cross-framework-impact")
def get_cross_framework_impact(finding_id: str):
    """Resolve the finding to its (framework, control_id) then ask the
    engine for the list of semantically equivalent controls in OTHER
    active frameworks. Cache-on-first-call, instant on subsequent calls.

    Returns an empty `mappings` list when the cache lookup + LLM call
    both yield nothing — the Registry UI shows a friendly empty state in
    that case.
    """
    finding = get_finding(finding_id)
    if not finding:
        raise HTTPException(404, detail="Finding not found")
    mappings = get_or_compute_mappings(
        finding["framework"], finding["control_id"])
    return {
        "finding_id":        finding_id,
        "source_framework":  finding["framework"],
        "source_control_id": finding["control_id"],
        "mappings":          mappings,
    }


# ─── Document Generation ─────────────────────────────────────────────

class GenerateRequest(BaseModel):
    framework_id: str
    doc_type: str
    context: dict = {}
    run_id: Optional[str] = None
    language: Optional[str] = "en"


@app.post("/api/v1/generate")
def generate(body: GenerateRequest):
    try:
        content = generate_document(body.doc_type, body.framework_id, body.context, body.language)
        doc_out_id = save_output_document(body.doc_type, body.framework_id,
                                           body.run_id, content)
        return {"document_out_id": doc_out_id, "doc_type": body.doc_type,
                "framework_id": body.framework_id, "content": content}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


class ExportRequest(BaseModel):
    content: str
    format: str  # "pdf" | "docx"
    filename: Optional[str] = "GRACE_Document"


@app.post("/api/v1/generate/export")
def export_document(body: ExportRequest):
    from modules.export import markdown_to_pdf_bytes, markdown_to_docx_bytes

    fmt = body.format.lower()
    if fmt == "pdf":
        try:
            data = markdown_to_pdf_bytes(body.content)
        except Exception as e:
            raise HTTPException(500, detail=f"PDF export failed: {e}")
        media_type = "application/pdf"
    elif fmt == "docx":
        try:
            data = markdown_to_docx_bytes(body.content)
        except Exception as e:
            raise HTTPException(500, detail=f"DOCX export failed: {e}")
        media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    else:
        raise HTTPException(400, detail="format must be 'pdf' or 'docx'")

    safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", body.filename or "GRACE_Document")
    return StreamingResponse(
        io.BytesIO(data),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.{fmt}"'},
    )


# ─── KPI / Dashboard ─────────────────────────────────────────────────

@app.get("/api/v1/kpi/summary")
def kpi_summary():
    return get_kpi_summary()


# ─── Risks ───────────────────────────────────────────────────────────

_RISK_CATEGORIES = {"operational", "cyber", "compliance", "financial",
                    "strategic", "reputational"}
_RISK_TREATMENTS = {"avoid", "transfer", "mitigate", "accept"}
_RISK_STATUSES = {"open", "under_treatment", "accepted", "closed"}


class RiskCreate(BaseModel):
    title: str
    description: Optional[str] = ""
    category: Optional[str] = "operational"
    likelihood: int = 3
    impact: int = 3
    residual_score: Optional[int] = None
    treatment_plan: Optional[str] = "mitigate"
    treatment_notes: Optional[str] = ""
    owner: Optional[str] = ""
    status: Optional[str] = "open"
    linked_controls: Optional[list] = None


class RiskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    likelihood: Optional[int] = None
    impact: Optional[int] = None
    residual_score: Optional[int] = None
    treatment_plan: Optional[str] = None
    treatment_notes: Optional[str] = None
    owner: Optional[str] = None
    status: Optional[str] = None
    linked_controls: Optional[list] = None


def _validate_risk_payload(d: dict):
    if "category" in d and d["category"] and d["category"] not in _RISK_CATEGORIES:
        raise HTTPException(400, detail=f"category must be one of {sorted(_RISK_CATEGORIES)}")
    if "treatment_plan" in d and d["treatment_plan"] and d["treatment_plan"] not in _RISK_TREATMENTS:
        raise HTTPException(400, detail=f"treatment_plan must be one of {sorted(_RISK_TREATMENTS)}")
    if "status" in d and d["status"] and d["status"] not in _RISK_STATUSES:
        raise HTTPException(400, detail=f"status must be one of {sorted(_RISK_STATUSES)}")
    for f in ("likelihood", "impact"):
        if f in d and d[f] is not None and not (1 <= int(d[f]) <= 5):
            raise HTTPException(400, detail=f"{f} must be between 1 and 5")


@app.post("/api/v1/risks", status_code=201)
def api_create_risk(body: RiskCreate):
    payload = body.model_dump(exclude_none=True)
    _validate_risk_payload(payload)
    if not payload.get("title", "").strip():
        raise HTTPException(400, detail="title is required")
    return create_risk(payload)


@app.get("/api/v1/risks")
def api_list_risks(status: Optional[str] = None, category: Optional[str] = None,
                    owner: Optional[str] = None, min_score: Optional[int] = None):
    return {"risks": list_risks(status=status, category=category,
                                 owner=owner, min_score=min_score)}


@app.get("/api/v1/risks/{risk_id}")
def api_get_risk(risk_id: str):
    r = get_risk(risk_id)
    if not r:
        raise HTTPException(404, detail="Risk not found")
    return r


@app.patch("/api/v1/risks/{risk_id}")
def api_update_risk(risk_id: str, body: RiskUpdate):
    payload = body.model_dump(exclude_none=True)
    _validate_risk_payload(payload)
    out = update_risk(risk_id, payload)
    if not out:
        raise HTTPException(404, detail="Risk not found")
    return out


@app.delete("/api/v1/risks/{risk_id}")
def api_delete_risk(risk_id: str):
    if not delete_risk(risk_id):
        raise HTTPException(404, detail="Risk not found")
    return {"status": "deleted", "risk_id": risk_id}


# ─── Vendors ─────────────────────────────────────────────────────────

_VENDOR_CATEGORIES = {"cloud_infra", "saas", "payment", "data_processor",
                      "professional_services", "other"}
_VENDOR_STATUSES = {"active", "under_review", "terminated"}
_VENDOR_TIERS = {"low", "medium", "high", "critical"}
_ANSWER_OPTIONS = {"yes", "no", "partial", "unknown"}


class VendorCreate(BaseModel):
    name: str
    category: Optional[str] = "saas"
    contact_email: Optional[str] = ""
    contract_url: Optional[str] = ""
    status: Optional[str] = "active"


class VendorUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    contact_email: Optional[str] = None
    contract_url: Optional[str] = None
    status: Optional[str] = None


class VendorAssessment(BaseModel):
    answers: list  # list of {question_id, question, answer, weight, notes}


@app.get("/api/v1/vendors/questionnaire-template")
def api_vendor_questionnaire_template():
    """Return the default 10-question template — frontend uses this to
    pre-populate the assessment form when no prior answers exist."""
    return {"questions": [
        {**q, "answer": "unknown", "notes": ""} for q in DEFAULT_VENDOR_QUESTIONS
    ]}


@app.post("/api/v1/vendors", status_code=201)
def api_create_vendor(body: VendorCreate):
    payload = body.model_dump(exclude_none=True)
    if not payload.get("name", "").strip():
        raise HTTPException(400, detail="name is required")
    if payload.get("category") and payload["category"] not in _VENDOR_CATEGORIES:
        raise HTTPException(400, detail=f"category must be one of {sorted(_VENDOR_CATEGORIES)}")
    if payload.get("status") and payload["status"] not in _VENDOR_STATUSES:
        raise HTTPException(400, detail=f"status must be one of {sorted(_VENDOR_STATUSES)}")
    return create_vendor(payload)


@app.get("/api/v1/vendors")
def api_list_vendors(risk_tier: Optional[str] = None,
                      category: Optional[str] = None,
                      status: Optional[str] = None):
    return {"vendors": list_vendors(risk_tier=risk_tier, category=category, status=status)}


@app.get("/api/v1/vendors/{vendor_id}")
def api_get_vendor(vendor_id: str):
    v = get_vendor(vendor_id)
    if not v:
        raise HTTPException(404, detail="Vendor not found")
    return v


@app.patch("/api/v1/vendors/{vendor_id}")
def api_update_vendor(vendor_id: str, body: VendorUpdate):
    payload = body.model_dump(exclude_none=True)
    if payload.get("category") and payload["category"] not in _VENDOR_CATEGORIES:
        raise HTTPException(400, detail=f"category must be one of {sorted(_VENDOR_CATEGORIES)}")
    if payload.get("status") and payload["status"] not in _VENDOR_STATUSES:
        raise HTTPException(400, detail=f"status must be one of {sorted(_VENDOR_STATUSES)}")
    out = update_vendor(vendor_id, payload)
    if not out:
        raise HTTPException(404, detail="Vendor not found")
    return out


@app.post("/api/v1/vendors/{vendor_id}/assess")
def api_assess_vendor(vendor_id: str, body: VendorAssessment):
    vendor = get_vendor(vendor_id)
    if not vendor:
        raise HTTPException(404, detail="Vendor not found")
    # Validate answer enum + normalise
    clean = []
    for a in body.answers or []:
        ans = (a.get("answer") or "unknown").lower()
        if ans not in _ANSWER_OPTIONS:
            ans = "unknown"
        clean.append({
            "question_id": a.get("question_id", ""),
            "question": a.get("question", ""),
            "answer": ans,
            "weight": int(a.get("weight", 0) or 0),
            "notes": a.get("notes", ""),
        })
    if not clean:
        raise HTTPException(400, detail="answers list cannot be empty")
    # Call Claude for AI summary (failure is non-fatal — saved as empty)
    try:
        summary = generate_vendor_assessment_summary(
            vendor.get("name", ""), vendor.get("category", ""), clean
        )
    except Exception as e:
        summary = f"_AI summary unavailable: {e}_"
    out = save_vendor_assessment(vendor_id, clean, ai_summary=summary)
    if not out:
        raise HTTPException(500, detail="Failed to save assessment")
    return out


@app.delete("/api/v1/vendors/{vendor_id}")
def api_delete_vendor(vendor_id: str):
    if not delete_vendor(vendor_id):
        raise HTTPException(404, detail="Vendor not found")
    return {"status": "deleted", "vendor_id": vendor_id}


# ─── Policies & Acknowledgments ──────────────────────────────────────

_POLICY_STATUSES = {"draft", "active", "superseded", "retired"}


class PolicyCreate(BaseModel):
    title: str
    version: str = "1.0"
    summary: Optional[str] = ""
    content: Optional[str] = ""
    effective_date: Optional[str] = ""
    review_date: Optional[str] = ""
    owner: Optional[str] = ""
    status: Optional[str] = "active"
    linked_controls: Optional[list] = None


class PolicyUpdate(BaseModel):
    title: Optional[str] = None
    version: Optional[str] = None
    summary: Optional[str] = None
    content: Optional[str] = None
    effective_date: Optional[str] = None
    review_date: Optional[str] = None
    owner: Optional[str] = None
    status: Optional[str] = None
    linked_controls: Optional[list] = None


class PolicyAssign(BaseModel):
    user_ids: list


class AcknowledgeRequest(BaseModel):
    signature_note: Optional[str] = ""


@app.post("/api/v1/policies", status_code=201)
def api_create_policy(body: PolicyCreate):
    payload = body.model_dump(exclude_none=True)
    if not payload.get("title", "").strip():
        raise HTTPException(400, detail="title is required")
    if payload.get("status") and payload["status"] not in _POLICY_STATUSES:
        raise HTTPException(400, detail=f"status must be one of {sorted(_POLICY_STATUSES)}")
    return create_policy(payload)


@app.get("/api/v1/policies")
def api_list_policies(status: Optional[str] = None):
    return {"policies": list_policies(status=status)}


@app.get("/api/v1/policies/{policy_id}")
def api_get_policy(policy_id: str):
    p = get_policy(policy_id)
    if not p:
        raise HTTPException(404, detail="Policy not found")
    return p


@app.patch("/api/v1/policies/{policy_id}")
def api_update_policy(policy_id: str, body: PolicyUpdate):
    payload = body.model_dump(exclude_none=True)
    if payload.get("status") and payload["status"] not in _POLICY_STATUSES:
        raise HTTPException(400, detail=f"status must be one of {sorted(_POLICY_STATUSES)}")
    out = update_policy(policy_id, payload)
    if not out:
        raise HTTPException(404, detail="Policy not found")
    return out


@app.post("/api/v1/policies/{policy_id}/assign")
def api_assign_policy(policy_id: str, body: PolicyAssign):
    if not get_policy(policy_id):
        raise HTTPException(404, detail="Policy not found")
    result = assign_policy(policy_id, body.user_ids or [])
    return result


@app.get("/api/v1/policy-assignments")
def api_list_assignments(user_id: Optional[str] = None,
                          status: Optional[str] = None,
                          policy_id: Optional[str] = None):
    return {"assignments": list_policy_assignments(
        user_id=user_id, status=status, policy_id=policy_id
    )}


@app.post("/api/v1/policy-assignments/{assignment_id}/acknowledge")
def api_acknowledge(assignment_id: str, body: AcknowledgeRequest):
    out = acknowledge_assignment(assignment_id, signature_note=body.signature_note or "")
    if not out:
        raise HTTPException(404, detail="Assignment not found")
    return out


# ─── Incidents ───────────────────────────────────────────────────────

_INCIDENT_SEVERITIES = {"low", "medium", "high", "critical"}
_INCIDENT_STATUSES = {"open", "investigating", "contained", "resolved", "closed"}
_INCIDENT_CATEGORIES = {"security_breach", "data_loss", "system_outage",
                        "policy_violation", "third_party", "other"}


class IncidentCreate(BaseModel):
    title: str
    description: Optional[str] = ""
    severity: str = "medium"
    status: Optional[str] = "open"
    category: Optional[str] = "other"
    reported_at: Optional[str] = None
    reported_by: Optional[str] = ""
    breach_notification_required: Optional[bool] = False
    impact_assessment: Optional[str] = ""
    root_cause: Optional[str] = ""
    remediation: Optional[str] = ""
    linked_controls: Optional[list] = None
    linked_findings: Optional[list] = None


class IncidentUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    severity: Optional[str] = None
    status: Optional[str] = None
    category: Optional[str] = None
    reported_by: Optional[str] = None
    resolved_at: Optional[str] = None
    breach_notification_required: Optional[bool] = None
    breach_notified_at: Optional[str] = None
    impact_assessment: Optional[str] = None
    root_cause: Optional[str] = None
    remediation: Optional[str] = None
    linked_controls: Optional[list] = None
    linked_findings: Optional[list] = None


def _validate_incident_payload(d: dict):
    if "severity" in d and d["severity"] and d["severity"] not in _INCIDENT_SEVERITIES:
        raise HTTPException(400, detail=f"severity must be one of {sorted(_INCIDENT_SEVERITIES)}")
    if "status" in d and d["status"] and d["status"] not in _INCIDENT_STATUSES:
        raise HTTPException(400, detail=f"status must be one of {sorted(_INCIDENT_STATUSES)}")
    if "category" in d and d["category"] and d["category"] not in _INCIDENT_CATEGORIES:
        raise HTTPException(400, detail=f"category must be one of {sorted(_INCIDENT_CATEGORIES)}")


@app.post("/api/v1/incidents", status_code=201)
def api_create_incident(body: IncidentCreate):
    payload = body.model_dump(exclude_none=True)
    _validate_incident_payload(payload)
    if not payload.get("title", "").strip():
        raise HTTPException(400, detail="title is required")
    return create_incident(payload)


@app.get("/api/v1/incidents")
def api_list_incidents(status: Optional[str] = None,
                        severity: Optional[str] = None,
                        category: Optional[str] = None):
    return {"incidents": list_incidents(status=status, severity=severity, category=category)}


@app.get("/api/v1/incidents/{incident_id}")
def api_get_incident(incident_id: str):
    inc = get_incident(incident_id)
    if not inc:
        raise HTTPException(404, detail="Incident not found")
    return inc


@app.patch("/api/v1/incidents/{incident_id}")
def api_update_incident(incident_id: str, body: IncidentUpdate):
    payload = body.model_dump(exclude_none=True)
    _validate_incident_payload(payload)
    out = update_incident(incident_id, payload)
    if not out:
        raise HTTPException(404, detail="Incident not found")
    return out


@app.delete("/api/v1/incidents/{incident_id}")
def api_delete_incident(incident_id: str):
    if not delete_incident(incident_id):
        raise HTTPException(404, detail="Incident not found")
    return {"status": "deleted", "incident_id": incident_id}
